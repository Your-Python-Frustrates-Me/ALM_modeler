"""
Deposit Elasticity Model for Interest Rate Risk
Модель эластичности депозитов для процентного риска

Моделирует изменение объемов депозитов в ответ на изменение процентных ставок.
Используется для построения динамического баланса в рамках процентного риска.
"""
from typing import Dict, List, Optional, Tuple
from datetime import date

from dataclasses import dataclass, field
import pandas as pd
import numpy as np
import logging
from enum import Enum

from alm_calculator.core.base_instrument import BaseInstrument, InstrumentType
from alm_calculator.models.instruments.deposit import Deposit

logger = logging.getLogger(__name__)


class CustomerSegment(str, Enum):
    """Сегменты клиентов для моделирования эластичности"""
    RETAIL = "retail"  # Физические лица (ФЛ)
    CORPORATE = "corporate"  # Юридические лица
    SME = "sme"  # Малый и средний бизнес
    GOVERNMENT = "government"  # Государственные организации
    BANK = "bank"  # Межбанковские депозиты


class DepositType(str, Enum):
    """Типы депозитов для моделирования эластичности"""
    DEMAND = "demand"  # До востребования
    SHORT_TERM = "short_term"  # До 3 месяцев
    MEDIUM_TERM = "medium_term"  # 3-12 месяцев
    LONG_TERM = "long_term"  # Свыше года


@dataclass
class ElasticityParameters:
    """
    Параметры эластичности для конкретного сегмента депозитов.

    Эластичность показывает процентное изменение объема депозитов
    при изменении ставки на 1 процентный пункт.

    Формула: ΔVolume% = elasticity × ΔRate%

    Пример:
        elasticity = -0.5 означает, что при росте ставки на 1% объем снизится на 0.5%
        (отток в другие банки или инструменты)
    """
    customer_segment: CustomerSegment
    deposit_type: DepositType

    # Базовая эластичность (линейная модель)
    base_elasticity: float = 0.0  # Обычно отрицательная для депозитов ФЛ

    # Параметры нелинейной модели
    use_nonlinear: bool = False  # Использовать нелинейную модель
    elasticity_ceiling: Optional[float] = None  # Максимальная эластичность (для больших шоков)
    elasticity_floor: Optional[float] = None  # Минимальная эластичность

    # Параметры пороговой модели (threshold model)
    threshold_rate_change: Optional[float] = None  # Порог изменения ставки (в п.п.)
    below_threshold_elasticity: Optional[float] = None  # Эластичность ниже порога
    above_threshold_elasticity: Optional[float] = None  # Эластичность выше порога

    # Параметры асимметричной модели
    asymmetric: bool = False  # Разная эластичность для роста и падения ставок
    positive_shock_elasticity: Optional[float] = None  # При росте ставок
    negative_shock_elasticity: Optional[float] = None  # При снижении ставок

    # Временные эффекты
    adjustment_speed: float = 1.0  # Скорость адаптации (0-1, где 1 = мгновенная)
    lag_days: int = 0  # Задержка реакции в днях

    # Конкурентные эффекты
    competitive_factor: float = 1.0  # Фактор конкурентности рынка (>1 = выше конкуренция)
    market_share_impact: bool = False  # Учитывать влияние рыночной доли

    # Ограничения
    max_volume_change: Optional[float] = None  # Максимальное изменение объема (доля)
    min_remaining_volume: Optional[float] = None  # Минимальный остаток (доля от исходного)

    @classmethod
    def create_retail_demand_default(cls) -> 'ElasticityParameters':
        """Дефолтные параметры для депозитов ФЛ до востребования"""
        return cls(
            customer_segment=CustomerSegment.RETAIL,
            deposit_type=DepositType.DEMAND,
            base_elasticity=-0.3,  # Умеренная чувствительность
            asymmetric=True,
            positive_shock_elasticity=-0.2,  # Слабая реакция на рост ставок (инерция)
            negative_shock_elasticity=-0.4,  # Более сильная на падение (отток)
            adjustment_speed=0.5,  # Медленная адаптация
            lag_days=30,  # Задержка месяц
            max_volume_change=0.15,  # Максимум 15% за период
            min_remaining_volume=0.60  # Минимум 60% остается (устойчивая часть)
        )

    @classmethod
    def create_retail_term_default(cls, deposit_type: DepositType = DepositType.SHORT_TERM) -> 'ElasticityParameters':
        """Дефолтные параметры для срочных депозитов ФЛ"""
        return cls(
            customer_segment=CustomerSegment.RETAIL,
            deposit_type=deposit_type,
            base_elasticity=-0.5,  # Выше чувствительность
            asymmetric=True,
            positive_shock_elasticity=-0.6,  # Сильная реакция на рост (переток)
            negative_shock_elasticity=-0.4,  # Чуть слабее на падение
            threshold_rate_change=1.0,  # Порог в 1 п.п.
            below_threshold_elasticity=-0.3,  # Слабая реакция на малые изменения
            above_threshold_elasticity=-0.8,  # Сильная на большие
            adjustment_speed=0.7,  # Быстрее чем demand
            lag_days=7,  # Неделя
            max_volume_change=0.25,  # До 25%
            min_remaining_volume=0.50
        )

    @classmethod
    def create_corporate_default(cls, deposit_type: DepositType = DepositType.SHORT_TERM) -> 'ElasticityParameters':
        """Дефолтные параметры для депозитов юридических лиц"""
        return cls(
            customer_segment=CustomerSegment.CORPORATE,
            deposit_type=deposit_type,
            base_elasticity=-0.8,  # Высокая чувствительность
            asymmetric=False,  # Симметричная реакция
            adjustment_speed=0.9,  # Быстрая адаптация
            lag_days=1,  # Практически мгновенно
            competitive_factor=1.5,  # Высокая конкуренция
            max_volume_change=0.40,  # До 40%
            min_remaining_volume=0.30
        )


@dataclass
class DepositVolumeChange:
    """Результат расчета изменения объема депозита"""
    instrument_id: str
    original_amount: float
    new_amount: float
    volume_change: float  # Абсолютное изменение
    volume_change_pct: float  # Процентное изменение
    rate_change_bps: float  # Изменение ставки в б.п.
    elasticity_used: float  # Примененная эластичность
    customer_segment: CustomerSegment
    deposit_type: DepositType


class DepositElasticityCalculator:
    """
    Калькулятор эластичности депозитов.

    Рассчитывает изменение объемов депозитов в ответ на изменение процентных ставок.
    Это отдельный расчет от текущего процентного риска, используется для
    построения динамического баланса.
    """

    def __init__(
        self,
        calculation_date: date,
        elasticity_params: Dict[Tuple[CustomerSegment, DepositType], ElasticityParameters]
    ):
        """
        Args:
            calculation_date: Дата расчета
            elasticity_params: Словарь параметров эластичности по сегментам
        """
        self.calculation_date = calculation_date
        self.elasticity_params = elasticity_params

    def calculate_volume_changes(
        self,
        deposits: List[Deposit],
        rate_shocks: Dict[str, float],  # {currency: shock_in_bps}
        customer_segment_mapper: Optional[callable] = None
    ) -> List[DepositVolumeChange]:
        """
        Рассчитывает изменение объемов депозитов для заданных шоков ставок.

        Args:
            deposits: Список депозитов
            rate_shocks: Словарь шоков ставок по валютам (в базисных пунктах)
            customer_segment_mapper: Функция для определения сегмента клиента
                                    (по умолчанию использует counterparty_type)

        Returns:
            Список изменений объемов депозитов
        """
        logger.info(
            f"Calculating deposit volume changes for {len(deposits)} deposits",
            extra={
                'calculation_date': str(self.calculation_date),
                'rate_shocks': rate_shocks,
                'deposits_count': len(deposits)
            }
        )

        volume_changes = []

        for deposit in deposits:
            # Получаем шок ставки для валюты депозита
            rate_shock_bps = rate_shocks.get(deposit.currency, 0.0)

            if abs(rate_shock_bps) < 0.01:
                # Нет изменения ставки - пропускаем
                continue

            # Определяем сегмент клиента
            segment = self._determine_customer_segment(deposit, customer_segment_mapper)

            # Определяем тип депозита
            deposit_type = self._determine_deposit_type(deposit)

            # Получаем параметры эластичности
            params_key = (segment, deposit_type)
            if params_key not in self.elasticity_params:
                # Пытаемся найти параметры для сегмента с любым типом депозита
                params_key = (segment, DepositType.DEMAND)
                if params_key not in self.elasticity_params:
                    logger.warning(
                        f"No elasticity parameters for {segment}/{deposit_type}, skipping",
                        extra={
                            'instrument_id': deposit.instrument_id,
                            'segment': segment.value,
                            'deposit_type': deposit_type.value
                        }
                    )
                    continue

            params = self.elasticity_params[params_key]

            # Рассчитываем изменение объема
            volume_change = self._calculate_single_deposit_change(
                deposit,
                rate_shock_bps,
                params
            )

            volume_changes.append(volume_change)

        logger.info(
            f"Calculated volume changes for {len(volume_changes)} deposits",
            extra={
                'total_original_volume': float(sum(vc.original_amount for vc in volume_changes)),
                'total_new_volume': float(sum(vc.new_amount for vc in volume_changes)),
                'total_change': float(sum(vc.volume_change for vc in volume_changes))
            }
        )

        return volume_changes

    def _calculate_single_deposit_change(
        self,
        deposit: Deposit,
        rate_shock_bps: float,
        params: ElasticityParameters
    ) -> DepositVolumeChange:
        """
        Рассчитывает изменение объема для одного депозита.

        Args:
            deposit: Депозит
            rate_shock_bps: Шок ставки в базисных пунктах
            params: Параметры эластичности

        Returns:
            DepositVolumeChange с результатами расчета
        """
        original_amount = deposit.amount
        rate_change_pct = rate_shock_bps / 100  # Конвертируем б.п. в проценты

        # Определяем применимую эластичность
        elasticity = self._determine_elasticity(rate_change_pct, params)

        # Базовый расчет изменения объема (в процентах)
        volume_change_pct = elasticity * rate_change_pct

        # Применяем скорость адаптации
        volume_change_pct *= params.adjustment_speed

        # Применяем конкурентный фактор
        volume_change_pct *= params.competitive_factor

        # Ограничение максимального изменения
        if params.max_volume_change:
            volume_change_pct = max(
                -params.max_volume_change,
                min(params.max_volume_change, volume_change_pct)
            )

        # Рассчитываем новый объем
        new_amount = original_amount * float(1 + volume_change_pct)

        # Ограничение минимального остатка
        if params.min_remaining_volume:
            min_amount = original_amount * float(params.min_remaining_volume)
            new_amount = max(min_amount, new_amount)

        volume_change = new_amount - original_amount

        segment = self._determine_customer_segment(deposit, None)
        deposit_type = self._determine_deposit_type(deposit)

        return DepositVolumeChange(
            instrument_id=deposit.instrument_id,
            original_amount=original_amount,
            new_amount=new_amount,
            volume_change=volume_change,
            volume_change_pct=float(volume_change / original_amount) if original_amount > 0 else 0.0,
            rate_change_bps=rate_shock_bps,
            elasticity_used=elasticity,
            customer_segment=segment,
            deposit_type=deposit_type
        )

    def _determine_elasticity(
        self,
        rate_change_pct: float,
        params: ElasticityParameters
    ) -> float:
        """
        Определяет применимую эластичность с учетом различных моделей.

        Args:
            rate_change_pct: Изменение ставки в процентах
            params: Параметры эластичности

        Returns:
            Эластичность для использования в расчетах
        """
        # Асимметричная модель
        if params.asymmetric:
            if rate_change_pct > 0 and params.positive_shock_elasticity is not None:
                elasticity = params.positive_shock_elasticity
            elif rate_change_pct < 0 and params.negative_shock_elasticity is not None:
                elasticity = params.negative_shock_elasticity
            else:
                elasticity = params.base_elasticity

        # Пороговая модель
        elif params.threshold_rate_change and params.below_threshold_elasticity and params.above_threshold_elasticity:
            if abs(rate_change_pct) < params.threshold_rate_change:
                elasticity = params.below_threshold_elasticity
            else:
                elasticity = params.above_threshold_elasticity

        # Линейная модель (по умолчанию)
        else:
            elasticity = params.base_elasticity

        # Применяем ограничения (ceiling/floor)
        if params.elasticity_ceiling is not None:
            elasticity = min(elasticity, params.elasticity_ceiling)

        if params.elasticity_floor is not None:
            elasticity = max(elasticity, params.elasticity_floor)

        return elasticity

    def _determine_customer_segment(
        self,
        deposit: Deposit,
        custom_mapper: Optional[callable]
    ) -> CustomerSegment:
        """Определяет сегмент клиента для депозита"""
        if custom_mapper:
            return custom_mapper(deposit)

        # Используем counterparty_type если доступен
        if hasattr(deposit, 'counterparty_type'):
            ctype = getattr(deposit, 'counterparty_type', '').lower()
            if 'retail' in ctype or 'физ' in ctype:
                return CustomerSegment.RETAIL
            elif 'corporate' in ctype or 'юр' in ctype:
                return CustomerSegment.CORPORATE
            elif 'sme' in ctype or 'мсб' in ctype:
                return CustomerSegment.SME
            elif 'gov' in ctype or 'государ' in ctype:
                return CustomerSegment.GOVERNMENT
            elif 'bank' in ctype or 'банк' in ctype:
                return CustomerSegment.BANK

        # Дефолт: розница
        return CustomerSegment.RETAIL

    def _determine_deposit_type(self, deposit: Deposit) -> DepositType:
        """Определяет тип депозита по его характеристикам"""
        if deposit.is_demand_deposit:
            return DepositType.DEMAND

        if deposit.maturity_date:
            days_to_maturity = (deposit.maturity_date - self.calculation_date).days

            if days_to_maturity <= 90:
                return DepositType.SHORT_TERM
            elif days_to_maturity <= 365:
                return DepositType.MEDIUM_TERM
            else:
                return DepositType.LONG_TERM

        # Дефолт
        return DepositType.DEMAND

    def create_dynamic_balance_sheet(
        self,
        deposits: List[Deposit],
        rate_shocks: Dict[str, float]
    ) -> Tuple[List[Deposit], pd.DataFrame]:
        """
        Создает динамический баланс с учетом эластичности депозитов.

        Args:
            deposits: Исходные депозиты
            rate_shocks: Шоки ставок по валютам

        Returns:
            Tuple[новые_депозиты, таблица_изменений]
        """
        # Рассчитываем изменения объемов
        volume_changes = self.calculate_volume_changes(deposits, rate_shocks)

        # Создаем словарь изменений
        changes_dict = {vc.instrument_id: vc for vc in volume_changes}

        # Создаем новые депозиты с обновленными объемами
        new_deposits = []
        for deposit in deposits:
            if deposit.instrument_id in changes_dict:
                change = changes_dict[deposit.instrument_id]
                # Копируем депозит и обновляем объем
                new_deposit = deposit.model_copy(deep=True)
                new_deposit.amount = change.new_amount
                new_deposits.append(new_deposit)
            else:
                # Депозит не изменился
                new_deposits.append(deposit.model_copy(deep=True))

        # Создаем таблицу изменений для анализа
        changes_df = pd.DataFrame([
            {
                'instrument_id': vc.instrument_id,
                'customer_segment': vc.customer_segment.value,
                'deposit_type': vc.deposit_type.value,
                'original_amount': float(vc.original_amount),
                'new_amount': float(vc.new_amount),
                'volume_change': float(vc.volume_change),
                'volume_change_pct': vc.volume_change_pct,
                'rate_change_bps': vc.rate_change_bps,
                'elasticity': vc.elasticity_used
            }
            for vc in volume_changes
        ])

        logger.info(
            "Created dynamic balance sheet with elasticity",
            extra={
                'deposits_count': len(new_deposits),
                'changed_deposits': len(volume_changes),
                'total_volume_change': float(sum(vc.volume_change for vc in volume_changes))
            }
        )

        return new_deposits, changes_df

    def analyze_elasticity_impact(
        self,
        volume_changes: List[DepositVolumeChange]
    ) -> pd.DataFrame:
        """
        Анализирует влияние эластичности по сегментам и типам депозитов.

        Args:
            volume_changes: Список изменений объемов

        Returns:
            DataFrame с агрегированным анализом по сегментам
        """
        if not volume_changes:
            return pd.DataFrame()

        data = []

        # Группируем по сегментам и типам
        for vc in volume_changes:
            data.append({
                'customer_segment': vc.customer_segment.value,
                'deposit_type': vc.deposit_type.value,
                'original_amount': float(vc.original_amount),
                'new_amount': float(vc.new_amount),
                'volume_change': float(vc.volume_change),
                'volume_change_pct': vc.volume_change_pct,
                'rate_change_bps': vc.rate_change_bps,
                'elasticity': vc.elasticity_used
            })

        df = pd.DataFrame(data)

        # Агрегируем по сегментам и типам
        summary = df.groupby(['customer_segment', 'deposit_type']).agg({
            'original_amount': 'sum',
            'new_amount': 'sum',
            'volume_change': 'sum',
            'volume_change_pct': 'mean',
            'elasticity': 'mean'
        }).reset_index()

        # Добавляем агрегированный процент изменения
        summary['agg_volume_change_pct'] = (
            (summary['new_amount'] - summary['original_amount']) / summary['original_amount']
        )

        return summary


def create_default_elasticity_config() -> Dict[Tuple[CustomerSegment, DepositType], ElasticityParameters]:
    """
    Создает дефолтную конфигурацию параметров эластичности.

    Returns:
        Словарь параметров эластичности для различных сегментов
    """
    config = {}

    # Физические лица (ФЛ)
    config[(CustomerSegment.RETAIL, DepositType.DEMAND)] = \
        ElasticityParameters.create_retail_demand_default()

    config[(CustomerSegment.RETAIL, DepositType.SHORT_TERM)] = \
        ElasticityParameters.create_retail_term_default(DepositType.SHORT_TERM)

    config[(CustomerSegment.RETAIL, DepositType.MEDIUM_TERM)] = \
        ElasticityParameters.create_retail_term_default(DepositType.MEDIUM_TERM)

    config[(CustomerSegment.RETAIL, DepositType.LONG_TERM)] = ElasticityParameters(
        customer_segment=CustomerSegment.RETAIL,
        deposit_type=DepositType.LONG_TERM,
        base_elasticity=-0.3,  # Ниже чувствительность для долгосрочных
        adjustment_speed=0.3,  # Медленная адаптация
        lag_days=90,
        max_volume_change=0.10,
        min_remaining_volume=0.70
    )

    # Юридические лица
    for dtype in [DepositType.DEMAND, DepositType.SHORT_TERM,
                  DepositType.MEDIUM_TERM, DepositType.LONG_TERM]:
        config[(CustomerSegment.CORPORATE, dtype)] = \
            ElasticityParameters.create_corporate_default(dtype)

    # МСБ - промежуточные параметры между ФЛ и ЮЛ
    config[(CustomerSegment.SME, DepositType.SHORT_TERM)] = ElasticityParameters(
        customer_segment=CustomerSegment.SME,
        deposit_type=DepositType.SHORT_TERM,
        base_elasticity=-0.6,
        adjustment_speed=0.8,
        lag_days=3,
        max_volume_change=0.30,
        min_remaining_volume=0.40
    )

    return config


def export_elasticity_results_to_excel(
    volume_changes: List[DepositVolumeChange],
    summary: pd.DataFrame,
    output_path: str
) -> None:
    """
    Экспортирует результаты расчета эластичности в Excel.

    Args:
        volume_changes: Список изменений объемов депозитов
        summary: Сводная таблица по сегментам
        output_path: Путь к выходному Excel файлу
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = openpyxl.Workbook()

    # Лист 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary['A1'] = "Deposit Elasticity Analysis - Summary"
    ws_summary['A1'].font = Font(size=14, bold=True)

    # Записываем сводную таблицу
    for r_idx, row in enumerate(dataframe_to_rows(summary, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 3:  # Header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            # Форматирование
            if c_idx in [3, 4, 5] and r_idx > 3:  # Суммы
                cell.number_format = '#,##0'
            elif c_idx in [6, 7, 8] and r_idx > 3:  # Проценты и эластичность
                cell.number_format = '0.00%'

    # Лист 2: Detailed Changes
    ws_details = wb.create_sheet(title="Detailed Changes")

    ws_details['A1'] = "Detailed Volume Changes"
    ws_details['A1'].font = Font(size=12, bold=True)

    headers = ['Instrument ID', 'Segment', 'Type', 'Original Amount', 'New Amount',
               'Change', 'Change %', 'Rate Shock (bps)', 'Elasticity']

    for c_idx, header in enumerate(headers, 1):
        cell = ws_details.cell(row=3, column=c_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for r_idx, vc in enumerate(volume_changes, 4):
        ws_details.cell(row=r_idx, column=1, value=vc.instrument_id)
        ws_details.cell(row=r_idx, column=2, value=vc.customer_segment.value)
        ws_details.cell(row=r_idx, column=3, value=vc.deposit_type.value)
        ws_details.cell(row=r_idx, column=4, value=float(vc.original_amount))
        ws_details.cell(row=r_idx, column=5, value=float(vc.new_amount))
        ws_details.cell(row=r_idx, column=6, value=float(vc.volume_change))
        ws_details.cell(row=r_idx, column=7, value=vc.volume_change_pct)
        ws_details.cell(row=r_idx, column=8, value=vc.rate_change_bps)
        ws_details.cell(row=r_idx, column=9, value=vc.elasticity_used)

        # Форматирование
        ws_details.cell(row=r_idx, column=4).number_format = '#,##0'
        ws_details.cell(row=r_idx, column=5).number_format = '#,##0'
        ws_details.cell(row=r_idx, column=6).number_format = '#,##0'
        ws_details.cell(row=r_idx, column=7).number_format = '0.00%'

    wb.save(output_path)
    logger.info(f"Elasticity results exported to {output_path}")
