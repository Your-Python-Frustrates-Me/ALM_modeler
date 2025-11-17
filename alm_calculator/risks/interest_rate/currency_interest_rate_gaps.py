"""
Currency Interest Rate Gaps Calculator
Расчет процентных гэпов по валютам

Процентные гэпы (Interest Rate Gaps) показывают дисбаланс между активами и пассивами,
чувствительными к изменению процентных ставок, в разрезе валют и временных периодов.
"""
from typing import List, Dict, Optional, Tuple
from datetime import date
from decimal import Decimal
import pandas as pd
import numpy as np
import logging

from alm_calculator.core.base_instrument import BaseInstrument

logger = logging.getLogger(__name__)


class CurrencyInterestRateGapCalculator:
    """
    Калькулятор процентных гэпов по валютам.

    Рассчитывает:
    1. Rate-sensitive assets (RSA) по временным бакетам и валютам
    2. Rate-sensitive liabilities (RSL) по временным бакетам и валютам
    3. Gap = RSA - RSL
    4. Cumulative gaps
    5. Gap ratio = Gap / Total Assets
    6. Sensitivity to parallel shift (NII impact, EVE impact)
    """

    def __init__(
        self,
        calculation_date: date,
        repricing_buckets: List[str],
        target_currencies: Optional[List[str]] = None
    ):
        """
        Args:
            calculation_date: Дата расчета
            repricing_buckets: Временные корзины для переоценки процентной ставки
            target_currencies: Список валют для анализа (None = все валюты)
        """
        self.calculation_date = calculation_date
        self.repricing_buckets = repricing_buckets
        self.target_currencies = target_currencies or ['RUB', 'USD', 'EUR', 'CNY']

    def calculate(
        self,
        instruments: List[BaseInstrument],
        risk_params: Dict
    ) -> Dict[str, pd.DataFrame]:
        """
        Рассчитывает процентные гэпы по валютам.

        Args:
            instruments: Список инструментов
            risk_params: Параметры расчета рисков

        Returns:
            Dict[currency, DataFrame] где DataFrame содержит:
            - bucket: временная корзина
            - rsa: rate-sensitive assets
            - rsl: rate-sensitive liabilities
            - gap: гэп (RSA - RSL)
            - cumulative_gap: кумулятивный гэп
            - gap_ratio: gap / total_assets
        """
        logger.info(
            f"Starting currency interest rate gap calculation",
            extra={
                'calculation_date': str(self.calculation_date),
                'currencies': self.target_currencies,
                'instruments_count': len(instruments)
            }
        )

        # Собираем repricing amounts по валютам
        repricing_by_currency = self._collect_repricing_by_currency(
            instruments,
            risk_params
        )

        # Рассчитываем гэпы для каждой валюты
        gaps_by_currency = {}

        for currency, repricing_data in repricing_by_currency.items():
            if currency not in self.target_currencies:
                continue

            gaps_df = self._calculate_gaps_for_currency(currency, repricing_data)
            gaps_by_currency[currency] = gaps_df

            logger.info(
                f"Calculated interest rate gaps for {currency}",
                extra={
                    'currency': currency,
                    'total_rsa': float(gaps_df['rsa'].sum()),
                    'total_rsl': float(gaps_df['rsl'].sum()),
                    'total_gap': float(gaps_df['gap'].sum()),
                    'final_cumulative_gap': float(gaps_df['cumulative_gap'].iloc[-1])
                }
            )

        return gaps_by_currency

    def calculate_sensitivity(
        self,
        gaps_by_currency: Dict[str, pd.DataFrame],
        rate_shock_bps: int = 100
    ) -> Dict[str, Dict]:
        """
        Рассчитывает чувствительность к изменению процентных ставок.

        Args:
            gaps_by_currency: Результат calculate()
            rate_shock_bps: Параллельный сдвиг кривой доходности (базисные пункты)

        Returns:
            Dict[currency, {
                'nii_impact_1y': Decimal,  # Влияние на NII за год
                'eve_impact': Decimal,      # Влияние на EVE (упрощенно)
                'gap_limits_breached': bool
            }]
        """
        sensitivity_by_currency = {}

        rate_shock = rate_shock_bps / 10000  # Конвертируем б.п. в десятичную дробь

        for currency, gaps_df in gaps_by_currency.items():
            # NII Impact (Net Interest Income) за 1 год
            # Упрощенная модель: сумма гэпов до 1 года * rate_shock
            buckets_1y = ['0-1m', '1-3m', '3-6m', '6-12m']
            gaps_1y = gaps_df[gaps_df['bucket'].isin(buckets_1y)]
            nii_impact = gaps_1y['gap'].sum() * rate_shock

            # EVE Impact (Economic Value of Equity)
            # Упрощенная модель: взвешиваем гэпы по duration бакетов
            eve_impact = self._calculate_eve_impact(gaps_df, rate_shock)

            # Проверка лимитов (пример: gap не должен превышать 20% активов)
            gap_limit_breached = any(abs(gaps_df['gap_ratio']) > 0.20)

            sensitivity_by_currency[currency] = {
                'nii_impact_1y': Decimal(str(nii_impact)),
                'eve_impact': Decimal(str(eve_impact)),
                'gap_limits_breached': gap_limit_breached,
                'rate_shock_bps': rate_shock_bps
            }

            logger.info(
                f"Sensitivity analysis for {currency} ({rate_shock_bps} bps shock)",
                extra={
                    'currency': currency,
                    'nii_impact': float(nii_impact),
                    'eve_impact': float(eve_impact),
                    'gap_limits_breached': gap_limit_breached
                }
            )

        return sensitivity_by_currency

    def _collect_repricing_by_currency(
        self,
        instruments: List[BaseInstrument],
        risk_params: Dict
    ) -> Dict[str, Dict[str, Dict[str, Decimal]]]:
        """
        Собирает repricing amounts по валютам и временным бакетам.

        Returns:
            Dict[currency, Dict[bucket, {'rsa': Decimal, 'rsl': Decimal}]]
        """
        repricing_data = {}

        for instrument in instruments:
            # Получаем risk contribution
            contribution = instrument.calculate_risk_contribution(
                self.calculation_date,
                risk_params
            )

            if not contribution.repricing_date:
                # Инструмент не чувствителен к процентным ставкам
                continue

            # Определяем bucket по repricing date
            bucket = self._date_to_bucket(contribution.repricing_date)
            if not bucket:
                continue

            currency = instrument.currency

            if currency not in repricing_data:
                repricing_data[currency] = {
                    bucket: {'rsa': Decimal(0), 'rsl': Decimal(0)}
                    for bucket in self.repricing_buckets
                }

            if bucket not in repricing_data[currency]:
                repricing_data[currency][bucket] = {'rsa': Decimal(0), 'rsl': Decimal(0)}

            # Распределяем по RSA/RSL
            if contribution.repricing_amount > 0:
                # Актив
                repricing_data[currency][bucket]['rsa'] += contribution.repricing_amount
            else:
                # Пассив
                repricing_data[currency][bucket]['rsl'] += abs(contribution.repricing_amount)

        return repricing_data

    def _date_to_bucket(self, repricing_date: date) -> Optional[str]:
        """
        Конвертирует дату переоценки в временной бакет.
        """
        days = (repricing_date - self.calculation_date).days

        if days < 0:
            return None

        # Маппинг дней на бакеты
        if days <= 30:
            return '0-1m'
        elif days <= 90:
            return '1-3m'
        elif days <= 180:
            return '3-6m'
        elif days <= 365:
            return '6-12m'
        elif days <= 730:
            return '1-2y'
        elif days <= 1095:
            return '2-3y'
        elif days <= 1825:
            return '3-5y'
        elif days <= 2555:
            return '5-7y'
        elif days <= 3650:
            return '7-10y'
        else:
            return '10y+'

    def _calculate_gaps_for_currency(
        self,
        currency: str,
        repricing_data: Dict[str, Dict[str, Decimal]]
    ) -> pd.DataFrame:
        """
        Рассчитывает процентные гэпы для одной валюты.

        Args:
            currency: Код валюты
            repricing_data: Словарь с RSA и RSL по бакетам

        Returns:
            DataFrame с гэпами
        """
        data = []
        total_assets = Decimal(0)

        # Сначала считаем total assets
        for bucket_data in repricing_data.values():
            total_assets += bucket_data['rsa']

        # Рассчитываем гэпы по бакетам
        for bucket in self.repricing_buckets:
            bucket_data = repricing_data.get(bucket, {'rsa': Decimal(0), 'rsl': Decimal(0)})

            rsa = float(bucket_data['rsa'])
            rsl = float(bucket_data['rsl'])
            gap = rsa - rsl

            # Gap ratio
            if total_assets > 0:
                gap_ratio = gap / float(total_assets)
            else:
                gap_ratio = 0.0

            data.append({
                'bucket': bucket,
                'rsa': rsa,
                'rsl': rsl,
                'gap': gap,
                'gap_ratio': gap_ratio
            })

        df = pd.DataFrame(data)

        # Рассчитываем кумулятивные гэпы
        df['cumulative_gap'] = df['gap'].cumsum()

        return df

    def _calculate_eve_impact(
        self,
        gaps_df: pd.DataFrame,
        rate_shock: float
    ) -> float:
        """
        Рассчитывает влияние на Economic Value of Equity (упрощенно).

        EVE Impact ≈ Σ (Gap_i × Duration_i × ΔR)

        Args:
            gaps_df: DataFrame с гэпами
            rate_shock: Изменение процентной ставки (десятичная дробь)

        Returns:
            EVE impact
        """
        # Упрощенная duration для каждого бакета (midpoint)
        bucket_durations = {
            '0-1m': 0.5 / 12,
            '1-3m': 2 / 12,
            '3-6m': 4.5 / 12,
            '6-12m': 9 / 12,
            '1-2y': 1.5,
            '2-3y': 2.5,
            '3-5y': 4,
            '5-7y': 6,
            '7-10y': 8.5,
            '10y+': 12
        }

        eve_impact = 0.0

        for _, row in gaps_df.iterrows():
            bucket = row['bucket']
            gap = row['gap']
            duration = bucket_durations.get(bucket, 1.0)

            # EVE impact для этого бакета
            eve_impact += gap * duration * rate_shock

        return -eve_impact  # Отрицательное, т.к. рост ставок снижает EVE


def export_to_excel(
    gaps_by_currency: Dict[str, pd.DataFrame],
    sensitivity: Dict[str, Dict],
    output_path: str
) -> None:
    """
    Экспортирует процентные гэпы в Excel.

    Args:
        gaps_by_currency: Результат CurrencyInterestRateGapCalculator.calculate()
        sensitivity: Результат calculate_sensitivity()
        output_path: Путь к выходному Excel файлу
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()

    # Лист 1: Summary & Sensitivity
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary['A1'] = "Interest Rate Gaps & Sensitivity Analysis"
    ws_summary['A1'].font = Font(size=14, bold=True)

    # Sensitivity table
    ws_summary['A3'] = "Sensitivity Analysis"
    ws_summary['A3'].font = Font(size=12, bold=True)

    headers = ['Currency', 'NII Impact (1Y)', 'EVE Impact', 'Rate Shock (bps)', 'Gap Limit Breached']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for row_idx, (currency, sens_data) in enumerate(sensitivity.items(), 5):
        ws_summary.cell(row=row_idx, column=1, value=currency)
        ws_summary.cell(row=row_idx, column=2, value=float(sens_data['nii_impact_1y']))
        ws_summary.cell(row=row_idx, column=3, value=float(sens_data['eve_impact']))
        ws_summary.cell(row=row_idx, column=4, value=sens_data['rate_shock_bps'])
        ws_summary.cell(row=row_idx, column=5, value='YES' if sens_data['gap_limits_breached'] else 'NO')

        # Подсветка если лимит превышен
        if sens_data['gap_limits_breached']:
            ws_summary.cell(row=row_idx, column=5).font = Font(color="FF0000", bold=True)

    # Лист для каждой валюты
    for currency, gaps_df in gaps_by_currency.items():
        ws = wb.create_sheet(title=currency)

        ws['A1'] = f"Interest Rate Gaps - {currency}"
        ws['A1'].font = Font(size=12, bold=True)

        # Таблица с гэпами
        for r_idx, row in enumerate(dataframe_to_rows(gaps_df, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                if r_idx == 3:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # Форматирование чисел
                if c_idx in [2, 3, 4, 6] and r_idx > 3:  # Колонки с суммами
                    cell.number_format = '#,##0'
                elif c_idx == 5 and r_idx > 3:  # Gap ratio
                    cell.number_format = '0.0%'

                # Подсветка больших гэпов
                if c_idx == 5 and r_idx > 3 and isinstance(cell.value, (int, float)) and abs(cell.value) > 0.20:
                    cell.font = Font(color="FF0000")

        # График гэпов
        chart = BarChart()
        chart.title = f"Interest Rate Gaps - {currency}"
        chart.y_axis.title = "Amount"
        chart.x_axis.title = "Time Bucket"

        data_start_row = 4
        data_end_row = 3 + len(gaps_df)

        # RSA, RSL
        rsa_ref = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
        rsl_ref = Reference(ws, min_col=3, min_row=data_start_row, max_row=data_end_row)
        categories = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)

        chart.add_data(rsa_ref, titles_from_data=False)
        chart.add_data(rsl_ref, titles_from_data=False)
        chart.set_categories(categories)

        ws.add_chart(chart, "H3")

    wb.save(output_path)
    logger.info(f"Currency interest rate gaps exported to {output_path}")
