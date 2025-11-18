"""
Dynamic Balance Sheet Interest Rate Risk Calculator
Калькулятор процентного риска на динамическом балансе с эластичностью депозитов

Этот модуль объединяет расчет эластичности депозитов и процентного риска,
создавая динамический баланс, где объемы депозитов меняются в ответ на изменение ставок.
"""
from typing import Dict, List, Optional, Tuple
from datetime import date

import pandas as pd
import logging

from alm_calculator.core.base_instrument import BaseInstrument, BookType
from alm_calculator.models.instruments.deposit import Deposit
from alm_calculator.risks.interest_rate.currency_interest_rate_gaps import (
    CurrencyInterestRateGapCalculator
)
from alm_calculator.risks.interest_rate.deposit_elasticity import (
    DepositElasticityCalculator,
    ElasticityParameters,
    CustomerSegment,
    DepositType,
    DepositVolumeChange,
    create_default_elasticity_config
)

logger = logging.getLogger(__name__)


class DynamicBalanceIRRCalculator:
    """
    Калькулятор процентного риска на динамическом балансе.

    Отличия от стандартного CurrencyInterestRateGapCalculator:
    1. Учитывает изменение объемов депозитов при изменении ставок (эластичность)
    2. Создает динамический баланс с обновленными объемами
    3. Рассчитывает процентный риск на этом динамическом балансе
    4. Предоставляет сравнение со статическим балансом

    Это отдельный расчет от текущего процентного риска.
    """

    def __init__(
        self,
        calculation_date: date,
        repricing_buckets: List[str],
        elasticity_params: Optional[Dict[Tuple[CustomerSegment, DepositType], ElasticityParameters]] = None,
        target_currencies: Optional[List[str]] = None
    ):
        """
        Args:
            calculation_date: Дата расчета
            repricing_buckets: Временные корзины для переоценки
            elasticity_params: Параметры эластичности (если None, используются дефолтные)
            target_currencies: Список валют для анализа
        """
        self.calculation_date = calculation_date
        self.repricing_buckets = repricing_buckets
        self.target_currencies = target_currencies or ['RUB', 'USD', 'EUR', 'CNY']

        # Используем дефолтные параметры эластичности если не заданы
        if elasticity_params is None:
            elasticity_params = create_default_elasticity_config()

        self.elasticity_params = elasticity_params

        # Создаем калькуляторы
        self.elasticity_calculator = DepositElasticityCalculator(
            calculation_date=calculation_date,
            elasticity_params=elasticity_params
        )

        self.gap_calculator = CurrencyInterestRateGapCalculator(
            calculation_date=calculation_date,
            repricing_buckets=repricing_buckets,
            target_currencies=target_currencies
        )

    def calculate_dynamic_irr(
        self,
        instruments: List[BaseInstrument],
        rate_shocks: Dict[str, float],  # {currency: shock_in_bps}
        risk_params: Dict,
        book_filter: Optional[BookType] = None
    ) -> Dict:
        """
        Рассчитывает процентный риск на динамическом балансе с эластичностью.

        Args:
            instruments: Список инструментов
            rate_shocks: Шоки ставок по валютам (в базисных пунктах)
            risk_params: Параметры расчета рисков
            book_filter: Фильтр по книге (TRADING/BANKING)

        Returns:
            Dict с результатами:
            {
                'static': {  # Статический баланс (без эластичности)
                    'gaps': Dict[currency, DataFrame],
                    'sensitivity': Dict[currency, Dict]
                },
                'dynamic': {  # Динамический баланс (с эластичностью)
                    'gaps': Dict[currency, DataFrame],
                    'sensitivity': Dict[currency, Dict],
                    'volume_changes': List[DepositVolumeChange],
                    'elasticity_summary': DataFrame
                },
                'comparison': {  # Сравнение
                    'gap_differences': Dict[currency, DataFrame],
                    'nii_impact_difference': Dict[currency, Decimal],
                    'eve_impact_difference': Dict[currency, Decimal]
                }
            }
        """
        logger.info(
            "Starting dynamic balance IRR calculation",
            extra={
                'calculation_date': str(self.calculation_date),
                'instruments_count': len(instruments),
                'rate_shocks': rate_shocks,
                'book_filter': book_filter.value if book_filter else 'all'
            }
        )

        # Фильтруем инструменты по книге если нужно
        if book_filter is not None:
            instruments = [inst for inst in instruments if inst.get_book() == book_filter]

        # 1. Рассчитываем статический баланс (без эластичности)
        logger.info("Calculating static balance IRR")
        static_gaps = self.gap_calculator.calculate(instruments, risk_params, book_filter=None)
        static_sensitivity = self.gap_calculator.calculate_sensitivity(
            static_gaps,
            rate_shock_bps=list(rate_shocks.values())[0] if rate_shocks else 100
        )

        # 2. Применяем эластичность к депозитам
        logger.info("Applying elasticity to deposits")
        deposits = [inst for inst in instruments if isinstance(inst, Deposit)]
        non_deposits = [inst for inst in instruments if not isinstance(inst, Deposit)]

        # Создаем динамический баланс
        dynamic_deposits, elasticity_changes_df = self.elasticity_calculator.create_dynamic_balance_sheet(
            deposits,
            rate_shocks
        )

        # Получаем детальную информацию об изменениях
        volume_changes = self.elasticity_calculator.calculate_volume_changes(
            deposits,
            rate_shocks
        )

        # Анализ влияния эластичности
        elasticity_summary = self.elasticity_calculator.analyze_elasticity_impact(volume_changes)

        # 3. Рассчитываем процентный риск на динамическом балансе
        logger.info("Calculating dynamic balance IRR")
        dynamic_instruments = non_deposits + dynamic_deposits
        dynamic_gaps = self.gap_calculator.calculate(dynamic_instruments, risk_params, book_filter=None)
        dynamic_sensitivity = self.gap_calculator.calculate_sensitivity(
            dynamic_gaps,
            rate_shock_bps=list(rate_shocks.values())[0] if rate_shocks else 100
        )

        # 4. Сравниваем результаты
        logger.info("Comparing static vs dynamic balance")
        comparison = self._compare_static_vs_dynamic(
            static_gaps,
            static_sensitivity,
            dynamic_gaps,
            dynamic_sensitivity
        )

        result = {
            'static': {
                'gaps': static_gaps,
                'sensitivity': static_sensitivity
            },
            'dynamic': {
                'gaps': dynamic_gaps,
                'sensitivity': dynamic_sensitivity,
                'volume_changes': volume_changes,
                'elasticity_summary': elasticity_summary,
                'elasticity_changes_df': elasticity_changes_df
            },
            'comparison': comparison
        }

        logger.info(
            "Dynamic balance IRR calculation completed",
            extra={
                'deposits_changed': len(volume_changes),
                'total_volume_change': float(sum(vc.volume_change for vc in volume_changes))
            }
        )

        return result

    def calculate_multiple_scenarios(
        self,
        instruments: List[BaseInstrument],
        scenarios: Dict[str, Dict[str, float]],  # {scenario_name: {currency: shock_bps}}
        risk_params: Dict,
        book_filter: Optional[BookType] = None
    ) -> Dict[str, Dict]:
        """
        Рассчитывает процентный риск с эластичностью для нескольких сценариев.

        Args:
            instruments: Список инструментов
            scenarios: Словарь сценариев {имя: {валюта: шок}}
            risk_params: Параметры расчета рисков
            book_filter: Фильтр по книге

        Returns:
            Dict[scenario_name, результат_calculate_dynamic_irr]
        """
        results = {}

        for scenario_name, rate_shocks in scenarios.items():
            logger.info(f"Calculating dynamic IRR for scenario: {scenario_name}")
            results[scenario_name] = self.calculate_dynamic_irr(
                instruments,
                rate_shocks,
                risk_params,
                book_filter
            )

        return results

    def _compare_static_vs_dynamic(
        self,
        static_gaps: Dict[str, pd.DataFrame],
        static_sensitivity: Dict[str, Dict],
        dynamic_gaps: Dict[str, pd.DataFrame],
        dynamic_sensitivity: Dict[str, Dict]
    ) -> Dict:
        """
        Сравнивает результаты статического и динамического баланса.

        Returns:
            Dict с различиями между статическим и динамическим балансом
        """
        comparison = {
            'gap_differences': {},
            'nii_impact_difference': {},
            'eve_impact_difference': {},
            'gap_ratio_changes': {}
        }

        for currency in static_gaps.keys():
            if currency not in dynamic_gaps:
                continue

            static_df = static_gaps[currency]
            dynamic_df = dynamic_gaps[currency]

            # Разница в гэпах
            diff_df = static_df.copy()
            diff_df['rsa_diff'] = dynamic_df['rsa'] - static_df['rsa']
            diff_df['rsl_diff'] = dynamic_df['rsl'] - static_df['rsl']
            diff_df['gap_diff'] = dynamic_df['gap'] - static_df['gap']
            diff_df['gap_ratio_diff'] = dynamic_df['gap_ratio'] - static_df['gap_ratio']

            comparison['gap_differences'][currency] = diff_df

            # Разница в чувствительности
            if currency in static_sensitivity and currency in dynamic_sensitivity:
                static_sens = static_sensitivity[currency]
                dynamic_sens = dynamic_sensitivity[currency]

                comparison['nii_impact_difference'][currency] = (
                    dynamic_sens['nii_impact_1y'] - static_sens['nii_impact_1y']
                )

                comparison['eve_impact_difference'][currency] = (
                    dynamic_sens['eve_impact'] - static_sens['eve_impact']
                )

                # Изменение gap ratio (среднее по бакетам)
                comparison['gap_ratio_changes'][currency] = float(
                    diff_df['gap_ratio_diff'].abs().mean()
                )

        return comparison


def export_dynamic_irr_to_excel(
    result: Dict,
    output_path: str,
    scenario_name: str = "Base Scenario"
) -> None:
    """
    Экспортирует результаты динамического процентного риска в Excel.

    Args:
        result: Результат DynamicBalanceIRRCalculator.calculate_dynamic_irr()
        output_path: Путь к выходному Excel файлу
        scenario_name: Название сценария
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()

    # === Лист 1: Executive Summary ===
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary['A1'] = f"Dynamic Balance IRR Analysis - {scenario_name}"
    ws_summary['A1'].font = Font(size=14, bold=True)

    # Elasticity Summary
    row = 3
    ws_summary[f'A{row}'] = "Deposit Volume Changes (Elasticity)"
    ws_summary[f'A{row}'].font = Font(size=12, bold=True)
    row += 1

    if 'elasticity_summary' in result['dynamic'] and not result['dynamic']['elasticity_summary'].empty:
        elasticity_df = result['dynamic']['elasticity_summary']

        for r_idx, row_data in enumerate(dataframe_to_rows(elasticity_df, index=False, header=True), row):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)

                if r_idx == row:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        row = r_idx + 2

    # Sensitivity Comparison
    ws_summary[f'A{row}'] = "NII Impact Comparison (Static vs Dynamic)"
    ws_summary[f'A{row}'].font = Font(size=12, bold=True)
    row += 2

    headers = ['Currency', 'Static NII Impact', 'Dynamic NII Impact', 'Difference', 'Difference %']
    for c_idx, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=c_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    row += 1

    for currency in result['static']['sensitivity'].keys():
        static_nii = result['static']['sensitivity'][currency]['nii_impact_1y']
        dynamic_nii = result['dynamic']['sensitivity'][currency]['nii_impact_1y']
        diff = result['comparison']['nii_impact_difference'].get(currency, 0.0)
        diff_pct = float(diff / static_nii * 100) if static_nii != 0 else 0.0

        ws_summary.cell(row=row, column=1, value=currency)
        ws_summary.cell(row=row, column=2, value=float(static_nii))
        ws_summary.cell(row=row, column=3, value=float(dynamic_nii))
        ws_summary.cell(row=row, column=4, value=float(diff))
        ws_summary.cell(row=row, column=5, value=diff_pct)

        # Форматирование
        ws_summary.cell(row=row, column=2).number_format = '#,##0'
        ws_summary.cell(row=row, column=3).number_format = '#,##0'
        ws_summary.cell(row=row, column=4).number_format = '#,##0'
        ws_summary.cell(row=row, column=5).number_format = '0.0%'

        # Подсветка значительных изменений
        if abs(diff_pct) > 10:
            ws_summary.cell(row=row, column=5).font = Font(color="FF0000", bold=True)

        row += 1

    # === Листы для каждой валюты: Static vs Dynamic ===
    for currency in result['static']['gaps'].keys():
        ws = wb.create_sheet(title=f"{currency} - Comparison")

        ws['A1'] = f"Interest Rate Gaps - {currency} (Static vs Dynamic)"
        ws['A1'].font = Font(size=12, bold=True)

        # Static
        ws['A3'] = "Static Balance"
        ws['A3'].font = Font(size=11, bold=True)

        static_df = result['static']['gaps'][currency]
        start_row = 4
        for r_idx, row_data in enumerate(dataframe_to_rows(static_df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Dynamic
        dynamic_start_row = start_row + len(static_df) + 3
        ws[f'A{dynamic_start_row-1}'] = "Dynamic Balance (with Elasticity)"
        ws[f'A{dynamic_start_row-1}'].font = Font(size=11, bold=True)

        dynamic_df = result['dynamic']['gaps'][currency]
        for r_idx, row_data in enumerate(dataframe_to_rows(dynamic_df, index=False, header=True), dynamic_start_row):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == dynamic_start_row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Differences
        diff_start_row = dynamic_start_row + len(dynamic_df) + 3
        ws[f'A{diff_start_row-1}'] = "Differences (Dynamic - Static)"
        ws[f'A{diff_start_row-1}'].font = Font(size=11, bold=True)

        diff_df = result['comparison']['gap_differences'][currency]
        for r_idx, row_data in enumerate(dataframe_to_rows(diff_df[['bucket', 'gap_diff', 'gap_ratio_diff']],
                                                           index=False, header=True), diff_start_row):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == diff_start_row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # === Лист с детальными изменениями депозитов ===
    if result['dynamic']['volume_changes']:
        ws_details = wb.create_sheet(title="Deposit Changes Detail")

        ws_details['A1'] = "Detailed Deposit Volume Changes"
        ws_details['A1'].font = Font(size=12, bold=True)

        headers = ['Instrument ID', 'Segment', 'Type', 'Original', 'New', 'Change', 'Change %', 'Elasticity']
        for c_idx, header in enumerate(headers, 1):
            cell = ws_details.cell(row=3, column=c_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        for r_idx, vc in enumerate(result['dynamic']['volume_changes'], 4):
            ws_details.cell(row=r_idx, column=1, value=vc.instrument_id)
            ws_details.cell(row=r_idx, column=2, value=vc.customer_segment.value)
            ws_details.cell(row=r_idx, column=3, value=vc.deposit_type.value)
            ws_details.cell(row=r_idx, column=4, value=float(vc.original_amount))
            ws_details.cell(row=r_idx, column=5, value=float(vc.new_amount))
            ws_details.cell(row=r_idx, column=6, value=float(vc.volume_change))
            ws_details.cell(row=r_idx, column=7, value=vc.volume_change_pct)
            ws_details.cell(row=r_idx, column=8, value=vc.elasticity_used)

    wb.save(output_path)
    logger.info(f"Dynamic balance IRR results exported to {output_path}")
