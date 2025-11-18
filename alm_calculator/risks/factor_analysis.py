"""
Factor Analysis for Risk Metrics
Факторный анализ для декомпозиции изменений в метриках риска

Разлагает изменения в метриках риска (горизонт выживания, процентный риск) на:
1. Эффект снижения срочности старых сделок (maturity roll-down / aging effect)
2. Эффект новых сделок (new deals effect)

Для каждого нового продукта показывает его индивидуальное влияние на метрики.
"""
from typing import List, Dict, Optional, Tuple, Callable, Any
from datetime import date, timedelta

import pandas as pd
import numpy as np
import logging
from copy import deepcopy

from alm_calculator.core.base_instrument import BaseInstrument

logger = logging.getLogger(__name__)


class FactorAnalyzer:
    """
    Анализатор изменений в метриках риска с декомпозицией на факторы.

    Методология:
    1. Загружаем портфель на момент t-1 (предыдущий период)
    2. Загружаем портфель на момент t (текущий период)
    3. Идентифицируем существующие и новые продукты
    4. "Старим" существующие продукты на период (t - t-1)
    5. Рассчитываем метрики для трех состояний:
       - Базовое: портфель на t-1
       - Постаревшие позиции: существующие продукты на момент t
       - Полный портфель: все продукты на момент t
    6. Декомпозируем:
       - Δ_aging = Metric(aged) - Metric(base)
       - Δ_new = Metric(full) - Metric(aged)
       - Total Δ = Δ_aging + Δ_new
    """

    def __init__(
        self,
        base_date: date,
        comparison_date: date
    ):
        """
        Args:
            base_date: Базовая дата (t-1)
            comparison_date: Дата сравнения (t)
        """
        self.base_date = base_date
        self.comparison_date = comparison_date
        self.days_elapsed = (comparison_date - base_date).days

        if self.days_elapsed <= 0:
            raise ValueError(
                f"comparison_date ({comparison_date}) must be after base_date ({base_date})"
            )

    def analyze(
        self,
        base_instruments: List[BaseInstrument],
        comparison_instruments: List[BaseInstrument],
        metric_calculator: Callable[[List[BaseInstrument], date], Any],
        metric_name: str = "Risk Metric"
    ) -> Dict:
        """
        Проводит факторный анализ изменений в метрике риска.

        Args:
            base_instruments: Инструменты на базовую дату (t-1)
            comparison_instruments: Инструменты на дату сравнения (t)
            metric_calculator: Функция для расчета метрики.
                               Принимает (instruments: List[BaseInstrument], calc_date: date)
                               Возвращает метрику (число или словарь)
            metric_name: Название метрики для отчетности

        Returns:
            Dict с результатами анализа:
            {
                'metric_name': str,
                'base_date': date,
                'comparison_date': date,
                'days_elapsed': int,

                # Метрики
                'metric_base': Any,          # Метрика на базовую дату
                'metric_aged': Any,          # Метрика для постаревших позиций
                'metric_full': Any,          # Метрика для полного портфеля

                # Декомпозиция
                'total_change': Any,         # Общее изменение
                'aging_effect': Any,         # Эффект старения
                'new_deals_effect': Any,     # Эффект новых сделок

                # Продукты
                'existing_products_count': int,
                'new_products_count': int,
                'new_products': List[str],   # ID новых продуктов

                # Детали по новым продуктам (если calculate_individual=True)
                'new_products_breakdown': Optional[List[Dict]]
            }
        """
        logger.info(
            f"Starting factor analysis for {metric_name}",
            extra={
                'base_date': str(self.base_date),
                'comparison_date': str(self.comparison_date),
                'days_elapsed': self.days_elapsed,
                'base_instruments_count': len(base_instruments),
                'comparison_instruments_count': len(comparison_instruments)
            }
        )

        # 1. Идентифицируем существующие и новые продукты
        base_ids = set(inst.instrument_id for inst in base_instruments)
        comparison_ids = set(inst.instrument_id for inst in comparison_instruments)

        existing_ids = base_ids & comparison_ids
        new_ids = comparison_ids - base_ids

        logger.info(
            f"Product identification complete",
            extra={
                'existing_count': len(existing_ids),
                'new_count': len(new_ids),
                'disappeared_count': len(base_ids - comparison_ids)
            }
        )

        # 2. Создаем "постаревшие" инструменты
        aged_instruments = self._age_instruments(
            base_instruments,
            existing_ids,
            self.days_elapsed
        )

        # 3. Фильтруем инструменты для каждого состояния
        # Полный портфель на момент t
        full_instruments = comparison_instruments

        # 4. Рассчитываем метрики для трех состояний
        logger.info(f"Calculating {metric_name} for base state (t-1)")
        metric_base = metric_calculator(base_instruments, self.base_date)

        logger.info(f"Calculating {metric_name} for aged state (existing products at t)")
        metric_aged = metric_calculator(aged_instruments, self.comparison_date)

        logger.info(f"Calculating {metric_name} for full state (all products at t)")
        metric_full = metric_calculator(full_instruments, self.comparison_date)

        # 5. Рассчитываем изменения
        aging_effect = self._calculate_delta(metric_aged, metric_base)
        new_deals_effect = self._calculate_delta(metric_full, metric_aged)
        total_change = self._calculate_delta(metric_full, metric_base)

        logger.info(
            f"Factor decomposition complete",
            extra={
                'total_change': self._format_metric(total_change),
                'aging_effect': self._format_metric(aging_effect),
                'new_deals_effect': self._format_metric(new_deals_effect)
            }
        )

        # 6. Формируем результат
        result = {
            'metric_name': metric_name,
            'base_date': self.base_date,
            'comparison_date': self.comparison_date,
            'days_elapsed': self.days_elapsed,

            'metric_base': metric_base,
            'metric_aged': metric_aged,
            'metric_full': metric_full,

            'total_change': total_change,
            'aging_effect': aging_effect,
            'new_deals_effect': new_deals_effect,

            'existing_products_count': len(existing_ids),
            'new_products_count': len(new_ids),
            'new_products': sorted(list(new_ids)),

            'new_products_breakdown': None  # Будет заполнено в analyze_individual_impact
        }

        return result

    def analyze_individual_impact(
        self,
        base_instruments: List[BaseInstrument],
        comparison_instruments: List[BaseInstrument],
        metric_calculator: Callable[[List[BaseInstrument], date], Any],
        metric_name: str = "Risk Metric",
        top_n: Optional[int] = None
    ) -> Dict:
        """
        Анализирует индивидуальное влияние каждого нового продукта на метрику.

        Args:
            base_instruments: Инструменты на базовую дату
            comparison_instruments: Инструменты на дату сравнения
            metric_calculator: Функция для расчета метрики
            metric_name: Название метрики
            top_n: Показать только топ N продуктов по влиянию (None = все)

        Returns:
            Dict с результатами анализа, включая breakdown по каждому новому продукту
        """
        # Сначала проводим обычный анализ
        result = self.analyze(
            base_instruments,
            comparison_instruments,
            metric_calculator,
            metric_name
        )

        # Получаем постаревшие инструменты
        base_ids = set(inst.instrument_id for inst in base_instruments)
        comparison_ids = set(inst.instrument_id for inst in comparison_instruments)
        existing_ids = base_ids & comparison_ids
        new_ids = comparison_ids - base_ids

        aged_instruments = self._age_instruments(
            base_instruments,
            existing_ids,
            self.days_elapsed
        )

        # Метрика для базового состояния (постаревшие позиции)
        metric_aged = metric_calculator(aged_instruments, self.comparison_date)

        # Анализируем влияние каждого нового продукта
        logger.info(
            f"Analyzing individual impact of {len(new_ids)} new products",
            extra={'new_products_count': len(new_ids)}
        )

        new_products_breakdown = []

        for new_id in new_ids:
            # Находим новый инструмент
            new_instrument = next(
                (inst for inst in comparison_instruments if inst.instrument_id == new_id),
                None
            )

            if not new_instrument:
                continue

            # Создаем временный портфель: постаревшие + этот новый продукт
            temp_instruments = aged_instruments + [new_instrument]

            # Рассчитываем метрику с добавлением этого продукта
            metric_with_product = metric_calculator(temp_instruments, self.comparison_date)

            # Влияние этого продукта
            impact = self._calculate_delta(metric_with_product, metric_aged)

            new_products_breakdown.append({
                'product_id': new_id,
                'product_type': new_instrument.instrument_type.value,
                'amount': float(new_instrument.amount),
                'currency': new_instrument.currency,
                'maturity_date': new_instrument.maturity_date,
                'impact': impact,
                'impact_formatted': self._format_metric(impact)
            })

        # Сортируем по влиянию (по абсолютной величине)
        new_products_breakdown = sorted(
            new_products_breakdown,
            key=lambda x: self._get_impact_magnitude(x['impact']),
            reverse=True
        )

        # Ограничиваем топ N, если указано
        if top_n:
            new_products_breakdown = new_products_breakdown[:top_n]

        result['new_products_breakdown'] = new_products_breakdown

        logger.info(
            f"Individual impact analysis complete",
            extra={
                'analyzed_products': len(new_products_breakdown),
                'top_n': top_n
            }
        )

        return result

    def _age_instruments(
        self,
        instruments: List[BaseInstrument],
        existing_ids: set,
        days: int
    ) -> List[BaseInstrument]:
        """
        "Старит" инструменты на указанное количество дней.

        Обновляет:
        - as_of_date -> comparison_date
        - Для инструментов с maturity_date: уменьшает remaining maturity
        - Для бессрочных инструментов: изменяет только as_of_date

        Args:
            instruments: Список инструментов
            existing_ids: ID инструментов, которые нужно "состарить"
            days: Количество дней

        Returns:
            Список "постаревших" инструментов
        """
        aged_instruments = []

        for inst in instruments:
            if inst.instrument_id not in existing_ids:
                # Этот инструмент исчез к моменту t, пропускаем
                continue

            # Создаем копию инструмента
            aged_inst = deepcopy(inst)

            # Обновляем as_of_date
            aged_inst.as_of_date = self.comparison_date

            # Для инструментов с определенным сроком погашения,
            # maturity_date остается тем же (просто стал ближе)
            # Никаких дополнительных изменений не требуется

            aged_instruments.append(aged_inst)

        return aged_instruments

    def _calculate_delta(self, metric_new: Any, metric_old: Any) -> Any:
        """
        Рассчитывает изменение метрики.

        Обрабатывает разные типы метрик:
        - Числа (int, float)
        - Словари (рекурсивно)
        - DataFrames (по ключевым столбцам)
        """
        # Числовые метрики
        if isinstance(metric_new, (int, float, np.number)):
            return metric_new - metric_old

        # Словари (например, {'RUB': 30, 'USD': 25})
        if isinstance(metric_new, dict) and isinstance(metric_old, dict):
            delta = {}
            all_keys = set(metric_new.keys()) | set(metric_old.keys())
            for key in all_keys:
                val_new = metric_new.get(key, 0)
                val_old = metric_old.get(key, 0)
                delta[key] = self._calculate_delta(val_new, val_old)
            return delta

        # DataFrame (просто вернем разницу, если требуется)
        if isinstance(metric_new, pd.DataFrame):
            # Для DataFrame просто возвращаем как есть
            # В реальном случае можно сделать более сложную логику
            return metric_new

        # Для других типов возвращаем None
        return None

    def _format_metric(self, metric: Any) -> str:
        """Форматирует метрику для отображения"""
        if metric is None:
            return "N/A"

        if isinstance(metric, (int, float, np.number)):
            return f"{metric:,.2f}"

        if isinstance(metric, dict):
            return ", ".join([f"{k}: {self._format_metric(v)}" for k, v in metric.items()])

        return str(metric)

    def _get_impact_magnitude(self, impact: Any) -> float:
        """Получает абсолютную величину влияния для сортировки"""
        if isinstance(impact, (int, float, np.number)):
            return abs(float(impact))

        if isinstance(impact, dict):
            # Для словаря берем сумму абсолютных величин
            return sum(abs(float(v)) for v in impact.values() if isinstance(v, (int, float, np.number)))

        return 0.0


def export_to_excel(
    analysis_results: Dict,
    output_path: str
) -> None:
    """
    Экспортирует результаты факторного анализа в Excel.

    Args:
        analysis_results: Результаты analyze() или analyze_individual_impact()
        output_path: Путь к выходному Excel файлу
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = openpyxl.Workbook()

    # Лист 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary['A1'] = f"Factor Analysis: {analysis_results['metric_name']}"
    ws_summary['A1'].font = Font(size=14, bold=True)

    # Даты
    ws_summary['A3'] = "Base Date (t-1):"
    ws_summary['B3'] = analysis_results['base_date'].strftime('%Y-%m-%d')

    ws_summary['A4'] = "Comparison Date (t):"
    ws_summary['B4'] = analysis_results['comparison_date'].strftime('%Y-%m-%d')

    ws_summary['A5'] = "Days Elapsed:"
    ws_summary['B5'] = analysis_results['days_elapsed']

    # Метрики
    ws_summary['A7'] = "Metrics:"
    ws_summary['A7'].font = Font(size=12, bold=True)

    ws_summary['A8'] = "Base (t-1):"
    ws_summary['B8'] = str(analysis_results['metric_base'])

    ws_summary['A9'] = "Aged Positions (at t):"
    ws_summary['B9'] = str(analysis_results['metric_aged'])

    ws_summary['A10'] = "Full Portfolio (at t):"
    ws_summary['B10'] = str(analysis_results['metric_full'])

    # Декомпозиция
    ws_summary['A12'] = "Factor Decomposition:"
    ws_summary['A12'].font = Font(size=12, bold=True)

    ws_summary['A13'] = "Total Change:"
    ws_summary['B13'] = str(analysis_results['total_change'])
    ws_summary['B13'].font = Font(bold=True)

    ws_summary['A14'] = "  - Aging Effect:"
    ws_summary['B14'] = str(analysis_results['aging_effect'])

    ws_summary['A15'] = "  - New Deals Effect:"
    ws_summary['B15'] = str(analysis_results['new_deals_effect'])

    # Продукты
    ws_summary['A17'] = "Products:"
    ws_summary['A17'].font = Font(size=12, bold=True)

    ws_summary['A18'] = "Existing Products:"
    ws_summary['B18'] = analysis_results['existing_products_count']

    ws_summary['A19'] = "New Products:"
    ws_summary['B19'] = analysis_results['new_products_count']

    # Лист 2: New Products Breakdown (если есть)
    if analysis_results.get('new_products_breakdown'):
        ws_breakdown = wb.create_sheet(title="New Products Impact")

        ws_breakdown['A1'] = "Individual Impact of New Products"
        ws_breakdown['A1'].font = Font(size=12, bold=True)

        # Заголовки
        headers = ['Product ID', 'Type', 'Amount', 'Currency', 'Maturity Date', 'Impact']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_breakdown.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Данные
        for row_idx, product in enumerate(analysis_results['new_products_breakdown'], 4):
            ws_breakdown.cell(row=row_idx, column=1, value=product['product_id'])
            ws_breakdown.cell(row=row_idx, column=2, value=product['product_type'])
            ws_breakdown.cell(row=row_idx, column=3, value=product['amount'])
            ws_breakdown.cell(row=row_idx, column=4, value=product['currency'])
            ws_breakdown.cell(row=row_idx, column=5, value=str(product['maturity_date']) if product['maturity_date'] else 'N/A')
            ws_breakdown.cell(row=row_idx, column=6, value=product['impact_formatted'])

            # Форматирование
            ws_breakdown.cell(row=row_idx, column=3).number_format = '#,##0.00'

    wb.save(output_path)
    logger.info(f"Factor analysis results exported to {output_path}")
