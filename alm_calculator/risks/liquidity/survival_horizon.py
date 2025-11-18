"""
Survival Horizon Calculator
Расчет горизонта выживания банка

Горизонт выживания (Survival Horizon) - количество дней, в течение которых банк
может функционировать при отсутствии новых источников фондирования, используя
только имеющиеся ликвидные активы и контрактные денежные потоки.
"""
from typing import List, Dict, Optional
from datetime import date, timedelta
from decimal import Decimal
import pandas as pd
import numpy as np
import logging

from alm_calculator.core.base_instrument import BaseInstrument, RiskContribution

logger = logging.getLogger(__name__)


class SurvivalHorizonCalculator:
    """
    Калькулятор горизонта выживания.

    Методология (по модели вашего кода):
    1. Собирает все денежные потоки по дням для трех сценариев: NAME, MARKET, COMBO
    2. Добавляет буфер ликвидности как нулевой день
    3. Рассчитывает кумулятивный gap ликвидности
    4. Определяет горизонт выживания как первый день, когда кумулятивная позиция <= 0
    5. Если горизонт < 0 или > max_horizon_days, устанавливает max_horizon_days

    Сценарии:
    - NAME: консервативный сценарий (именной/name-based)
    - MARKET: рыночный сценарий
    - COMBO: комбинированный сценарий (обычно самый стрессовый)
    """

    def __init__(
        self,
        calculation_date: date,
        max_horizon_days: int = 90,
        scenarios: Optional[List[str]] = None
    ):
        """
        Args:
            calculation_date: Дата расчета
            max_horizon_days: Максимальный горизонт выживания в днях (по умолчанию 90)
            scenarios: Список сценариев для расчета (по умолчанию ['NAME', 'MARKET', 'COMBO'])
        """
        self.calculation_date = calculation_date
        self.max_horizon_days = max_horizon_days
        self.scenarios = scenarios or ['NAME', 'MARKET', 'COMBO']

    def calculate(
        self,
        daily_flows: pd.DataFrame,
        buffer: Dict[str, float],
        exclude_from_buffer: Optional[List[str]] = None
    ) -> Dict:
        """
        Рассчитывает горизонт выживания по дневным потокам.

        Args:
            daily_flows: DataFrame с дневными потоками по сценариям
                Обязательные колонки: FLOW_DAY (int), NAME, MARKET, COMBO (все в рублях)
                Опциональные: IN_BUFFER (флаг для исключения из расчета горизонта)
            buffer: Словарь с буфером ликвидности:
                {
                    'VALUE': float,  # Полная стоимость буфера
                    'IMPAIRMENT': float  # Обесценение (опционально)
                }
            exclude_from_buffer: Список названий сценариев для исключения потоков
                с флагом IN_BUFFER=1 (по умолчанию None - не исключаем)

        Returns:
            Словарь с результатами:
            {
                'horizon_days': {
                    'NAME': int,
                    'MARKET': int,
                    'COMBO': int
                },
                'cumulative_report': pd.DataFrame,  # Отчет с кумулятивными позициями
                'calculation_date': date,
                'buffer_value': float,
                'buffer_impaired_value': float
            }
        """
        logger.info(
            f"Starting survival horizon calculation",
            extra={
                'calculation_date': str(self.calculation_date),
                'scenarios': self.scenarios,
                'flows_count': len(daily_flows)
            }
        )

        # Фильтруем потоки (исключаем IN_BUFFER если нужно)
        if exclude_from_buffer and 'IN_BUFFER' in daily_flows.columns:
            flows_for_horizon = daily_flows[daily_flows['IN_BUFFER'] == 0].copy()
        else:
            flows_for_horizon = daily_flows.copy()

        # Группируем потоки по дням и сценариям
        scenario_columns = [s for s in self.scenarios if s in flows_for_horizon.columns]

        grouped_flows = flows_for_horizon.groupby('FLOW_DAY')[scenario_columns].sum()
        grouped_flows = grouped_flows.sort_index()

        # Определяем буфер
        buffer_value = buffer.get('VALUE', 0.0)
        buffer_impairment = buffer.get('IMPAIRMENT', 0.0)
        buffer_impaired_value = buffer_value - buffer_impairment

        # Создаем строку с буфером (день 0)
        buffer_row = pd.DataFrame(
            [[buffer_value] * len(scenario_columns)],
            columns=scenario_columns,
            index=[0]
        )

        # Для консервативных сценариев используем обесцененную стоимость
        if 'MARKET' in scenario_columns:
            buffer_row.at[0, 'MARKET'] = buffer_impaired_value
        if 'COMBO' in scenario_columns:
            buffer_row.at[0, 'COMBO'] = buffer_impaired_value

        # Добавляем буфер к потокам
        report = pd.concat([buffer_row, grouped_flows], ignore_index=False)
        report = report.fillna(0)

        # Считаем кумулятивную сумму
        cumsum_report = report.cumsum()
        cumsum_report['FLOW_DAY'] = np.arange(len(cumsum_report))
        cumsum_report['FLOW_DATE'] = [
            self.calculation_date + timedelta(days=int(day))
            for day in cumsum_report['FLOW_DAY']
        ]
        cumsum_report['REPORT_DATE'] = self.calculation_date

        # Рассчитываем горизонт выживания для каждого сценария
        horizon_days = {}

        for scenario in scenario_columns:
            # Находим первый день, когда кумулятивная позиция становится <= 0
            # np.argmin находит первый индекс, где условие True
            horizon = np.argmin(cumsum_report[scenario].values > 0) - 1

            # Применяем ограничения
            if horizon < 0 or horizon > self.max_horizon_days:
                horizon = self.max_horizon_days

            horizon_days[scenario] = int(horizon)

            logger.info(
                f"Survival horizon for {scenario}: {horizon} days",
                extra={
                    'scenario': scenario,
                    'survival_days': int(horizon),
                    'buffer': buffer_value
                }
            )

        return {
            'horizon_days': horizon_days,
            'cumulative_report': cumsum_report,
            'calculation_date': self.calculation_date,
            'buffer_value': buffer_value,
            'buffer_impaired_value': buffer_impaired_value
        }



def export_to_excel(results: Dict, output_path: str) -> None:
    """
    Экспортирует результаты в Excel для анализа.

    Args:
        results: Результаты расчета от SurvivalHorizonCalculator.calculate()
            {
                'horizon_days': {'NAME': int, 'MARKET': int, 'COMBO': int},
                'cumulative_report': pd.DataFrame,
                'calculation_date': date,
                'buffer_value': float,
                'buffer_impaired_value': float
            }
        output_path: Путь к выходному Excel файлу
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = openpyxl.Workbook()

    # Лист 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary['A1'] = "Survival Horizon Analysis"
    ws_summary['A1'].font = Font(size=14, bold=True)

    ws_summary['A3'] = "Calculation Date:"
    ws_summary['B3'] = results['calculation_date'].strftime('%Y-%m-%d')

    ws_summary['A4'] = "Liquid Assets Buffer:"
    ws_summary['B4'] = results['buffer_value']

    ws_summary['A5'] = "Buffer (impaired):"
    ws_summary['B5'] = results['buffer_impaired_value']

    # Горизонты выживания по сценариям
    ws_summary['A7'] = "Survival Horizon (days) by Scenario:"
    ws_summary['A7'].font = Font(bold=True)

    row_idx = 8
    min_horizon = min(results['horizon_days'].values())

    for scenario, horizon in results['horizon_days'].items():
        ws_summary[f'A{row_idx}'] = f"{scenario}:"
        ws_summary[f'B{row_idx}'] = horizon

        # Выделяем красным если < 30 дней
        if horizon < 30:
            ws_summary[f'B{row_idx}'].font = Font(bold=True, color="FF0000")
        elif horizon == min_horizon:
            ws_summary[f'B{row_idx}'].font = Font(bold=True)

        row_idx += 1

    # Лист 2: Cumulative Report
    ws_report = wb.create_sheet(title="Cumulative Report")

    ws_report['A1'] = "Cumulative Liquidity Positions"
    ws_report['A1'].font = Font(size=12, bold=True)

    # Таблица с кумулятивными позициями
    cumulative_report = results['cumulative_report']
    for r_idx, row in enumerate(dataframe_to_rows(cumulative_report, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws_report.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 3:  # Header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    wb.save(output_path)
    logger.info(f"Survival horizon results exported to {output_path}")
