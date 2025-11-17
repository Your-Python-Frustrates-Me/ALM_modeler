"""
Currency Liquidity Gaps Calculator
Расчет гэпов ликвидности по валютам

Гэпы ликвидности показывают дисбаланс между притоками и оттоками денежных средств
в разрезе валют и временных периодов.
"""
from typing import List, Dict, Optional
from datetime import date
from decimal import Decimal
import pandas as pd
import logging

from alm_calculator.core.base_instrument import BaseInstrument

logger = logging.getLogger(__name__)


class CurrencyLiquidityGapCalculator:
    """
    Калькулятор гэпов ликвидности по валютам.

    Рассчитывает:
    1. Притоки (inflows) по временным бакетам и валютам
    2. Оттоки (outflows) по временным бакетам и валютам
    3. Чистые гэпы (net gaps) = inflows - outflows
    4. Кумулятивные гэпы (cumulative gaps)
    5. Coverage ratios (отношение inflows к outflows)
    """

    def __init__(
        self,
        calculation_date: date,
        liquidity_buckets: List[str],
        target_currencies: Optional[List[str]] = None
    ):
        """
        Args:
            calculation_date: Дата расчета
            liquidity_buckets: Временные корзины для анализа ликвидности
            target_currencies: Список валют для анализа (None = все валюты)
        """
        self.calculation_date = calculation_date
        self.liquidity_buckets = liquidity_buckets
        self.target_currencies = target_currencies or ['RUB', 'USD', 'EUR', 'CNY']

    def calculate(
        self,
        instruments: List[BaseInstrument],
        risk_params: Dict
    ) -> Dict[str, pd.DataFrame]:
        """
        Рассчитывает гэпы ликвидности по валютам.

        Args:
            instruments: Список инструментов
            risk_params: Параметры расчета рисков

        Returns:
            Dict[currency, DataFrame] где DataFrame содержит:
            - bucket: временная корзина
            - inflow: притоки
            - outflow: оттоки
            - net_gap: чистый гэп
            - cumulative_gap: кумулятивный гэп
            - coverage_ratio: коэффициент покрытия (inflow/outflow)
        """
        logger.info(
            f"Starting currency liquidity gap calculation",
            extra={
                'calculation_date': str(self.calculation_date),
                'currencies': self.target_currencies,
                'instruments_count': len(instruments)
            }
        )

        # Собираем cash flows по валютам
        cash_flows_by_currency = self._collect_cash_flows_by_currency(
            instruments,
            risk_params
        )

        # Рассчитываем гэпы для каждой валюты
        gaps_by_currency = {}

        for currency, cash_flows in cash_flows_by_currency.items():
            if currency not in self.target_currencies:
                continue

            gaps_df = self._calculate_gaps_for_currency(currency, cash_flows)
            gaps_by_currency[currency] = gaps_df

            logger.info(
                f"Calculated liquidity gaps for {currency}",
                extra={
                    'currency': currency,
                    'total_inflow': float(gaps_df['inflow'].sum()),
                    'total_outflow': float(gaps_df['outflow'].sum()),
                    'net_gap': float(gaps_df['net_gap'].sum()),
                    'final_cumulative_gap': float(gaps_df['cumulative_gap'].iloc[-1])
                }
            )

        return gaps_by_currency

    def _collect_cash_flows_by_currency(
        self,
        instruments: List[BaseInstrument],
        risk_params: Dict
    ) -> Dict[str, Dict[str, Dict[str, Decimal]]]:
        """
        Собирает cash flows по валютам и временным бакетам.

        Returns:
            Dict[currency, Dict[bucket, {'inflow': Decimal, 'outflow': Decimal}]]
        """
        cash_flows = {}

        for instrument in instruments:
            # Получаем risk contribution
            contribution = instrument.calculate_risk_contribution(
                self.calculation_date,
                risk_params
            )

            # Определяем валюту инструмента
            currency = instrument.currency

            if currency not in cash_flows:
                cash_flows[currency] = {
                    bucket: {'inflow': Decimal(0), 'outflow': Decimal(0)}
                    for bucket in self.liquidity_buckets
                }

            # Распределяем cash flows по бакетам
            for bucket, cf_amount in contribution.cash_flows.items():
                if bucket not in cash_flows[currency]:
                    cash_flows[currency][bucket] = {'inflow': Decimal(0), 'outflow': Decimal(0)}

                if cf_amount > 0:
                    cash_flows[currency][bucket]['inflow'] += cf_amount
                else:
                    cash_flows[currency][bucket]['outflow'] += abs(cf_amount)

        return cash_flows

    def _calculate_gaps_for_currency(
        self,
        currency: str,
        cash_flows: Dict[str, Dict[str, Decimal]]
    ) -> pd.DataFrame:
        """
        Рассчитывает гэпы ликвидности для одной валюты.

        Args:
            currency: Код валюты
            cash_flows: Словарь с притоками и оттоками по бакетам

        Returns:
            DataFrame с гэпами
        """
        data = []

        for bucket in self.liquidity_buckets:
            bucket_data = cash_flows.get(bucket, {'inflow': Decimal(0), 'outflow': Decimal(0)})

            inflow = float(bucket_data['inflow'])
            outflow = float(bucket_data['outflow'])
            net_gap = inflow - outflow

            # Coverage ratio (коэффициент покрытия)
            if outflow > 0:
                coverage_ratio = inflow / outflow
            else:
                coverage_ratio = float('inf') if inflow > 0 else 1.0

            data.append({
                'bucket': bucket,
                'inflow': inflow,
                'outflow': outflow,
                'net_gap': net_gap,
                'coverage_ratio': coverage_ratio
            })

        df = pd.DataFrame(data)

        # Рассчитываем кумулятивные гэпы
        df['cumulative_gap'] = df['net_gap'].cumsum()

        return df

    def analyze_gaps(
        self,
        gaps_by_currency: Dict[str, pd.DataFrame]
    ) -> Dict:
        """
        Анализирует рассчитанные гэпы и выявляет проблемные зоны.

        Args:
            gaps_by_currency: Результат calculate()

        Returns:
            Словарь с аналитикой:
            {
                'critical_currencies': List[str],  # Валюты с отрицательными кум. гэпами
                'coverage_analysis': Dict,         # Анализ коэффициентов покрытия
                'alerts': List[str]                # Предупреждения
            }
        """
        critical_currencies = []
        coverage_analysis = {}
        alerts = []

        for currency, gaps_df in gaps_by_currency.items():
            # Проверяем кумулятивные гэпы
            min_cumulative_gap = gaps_df['cumulative_gap'].min()

            if min_cumulative_gap < 0:
                critical_currencies.append(currency)
                critical_bucket = gaps_df[gaps_df['cumulative_gap'] == min_cumulative_gap]['bucket'].iloc[0]
                alerts.append(
                    f"{currency}: Negative cumulative gap {min_cumulative_gap:,.0f} in bucket {critical_bucket}"
                )

            # Анализ coverage ratios
            low_coverage_buckets = gaps_df[gaps_df['coverage_ratio'] < 1.0]
            if not low_coverage_buckets.empty:
                coverage_analysis[currency] = {
                    'buckets_with_low_coverage': low_coverage_buckets['bucket'].tolist(),
                    'min_coverage_ratio': float(low_coverage_buckets['coverage_ratio'].min())
                }

                if low_coverage_buckets['coverage_ratio'].min() < 0.8:
                    alerts.append(
                        f"{currency}: Critical coverage ratio "
                        f"{low_coverage_buckets['coverage_ratio'].min():.2f} "
                        f"in buckets {', '.join(low_coverage_buckets['bucket'].tolist())}"
                    )

        return {
            'critical_currencies': critical_currencies,
            'coverage_analysis': coverage_analysis,
            'alerts': alerts
        }


def export_to_excel(
    gaps_by_currency: Dict[str, pd.DataFrame],
    analysis: Dict,
    output_path: str
) -> None:
    """
    Экспортирует гэпы ликвидности в Excel.

    Args:
        gaps_by_currency: Результат CurrencyLiquidityGapCalculator.calculate()
        analysis: Результат CurrencyLiquidityGapCalculator.analyze_gaps()
        output_path: Путь к выходному Excel файлу
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, Reference

    wb = openpyxl.Workbook()

    # Лист 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary['A1'] = "Currency Liquidity Gaps Analysis"
    ws_summary['A1'].font = Font(size=14, bold=True)

    # Alerts
    ws_summary['A3'] = "Alerts:"
    ws_summary['A3'].font = Font(bold=True, color="FF0000")

    if analysis['alerts']:
        for idx, alert in enumerate(analysis['alerts'], start=4):
            ws_summary[f'A{idx}'] = alert
            ws_summary[f'A{idx}'].font = Font(color="FF0000")
    else:
        ws_summary['A4'] = "No critical alerts"
        ws_summary['A4'].font = Font(color="00AA00")

    # Critical currencies
    row = 6 + len(analysis['alerts'])
    ws_summary[f'A{row}'] = "Critical Currencies:"
    ws_summary[f'A{row}'].font = Font(bold=True)

    if analysis['critical_currencies']:
        ws_summary[f'B{row}'] = ', '.join(analysis['critical_currencies'])
    else:
        ws_summary[f'B{row}'] = "None"

    # Лист для каждой валюты
    for currency, gaps_df in gaps_by_currency.items():
        ws = wb.create_sheet(title=currency)

        ws['A1'] = f"Liquidity Gaps - {currency}"
        ws['A1'].font = Font(size=12, bold=True)

        # Таблица с гэпами
        for r_idx, row in enumerate(dataframe_to_rows(gaps_df, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                if r_idx == 3:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # Форматирование чисел
                if c_idx in [2, 3, 4, 5] and r_idx > 3:  # Колонки с числами
                    cell.number_format = '#,##0'

                # Подсветка отрицательных значений
                if c_idx in [4, 5] and r_idx > 3 and isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.font = Font(color="FF0000")

        # График
        chart = LineChart()
        chart.title = f"Cumulative Gap - {currency}"
        chart.y_axis.title = "Amount"
        chart.x_axis.title = "Time Bucket"

        # Данные для графика
        data_start_row = 4
        data_end_row = 3 + len(gaps_df)

        # Cumulative gap
        cumulative_ref = Reference(ws, min_col=6, min_row=data_start_row, max_row=data_end_row)
        categories = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)

        chart.add_data(cumulative_ref, titles_from_data=False)
        chart.set_categories(categories)

        ws.add_chart(chart, f"H3")

    wb.save(output_path)
    logger.info(f"Currency liquidity gaps exported to {output_path}")
